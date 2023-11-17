import json
import requests
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render,redirect
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from .forms import ExcelFileForm
from .forms import CustomAuthenticationForm
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
import logging 
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.urls import reverse_lazy
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
def upload_excel(request):
    if request.method == 'POST':
        form = ExcelFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data['file']

            # Check if the file extension is valid
            if not uploaded_file.name.endswith('.xlsx'):
                return HttpResponseBadRequest('Invalid file format. Please upload an XLSX file.')

            try:
                workbook = load_workbook(uploaded_file, read_only=True)
                worksheet = workbook.active
                data = []
                for row in worksheet.iter_rows():
                    data.append([cell.value for cell in row])

                # Assuming you have an upload_excel.html template
                return render(request, 'display_excel.html', {'data': data, 'form': form, 'uploaded_file': uploaded_file})

            except InvalidFileException:
                return HttpResponseBadRequest('Invalid Excel file format')
        else:
            return render(request, 'upload_excel.html', {'form': form})
    else:
        form = ExcelFileForm()
    return render(request, 'upload_excel.html', {'form': form})

def save_excel_data(request):
    if request.method == "POST":
        if "save_button" in request.POST:
            # Extract headers from the first row
            headers = [cell for cell in request.POST.getlist("headers[]")]

            # Extract data from the remaining rows
            data = request.POST.getlist("data[]")

            # Calculate the number of columns based on the headers
            num_columns = len(headers)

            # Create a list of dictionaries where each dictionary represents a row
            rows = [dict(zip(headers, data[i:i+num_columns])) for i in range(0, len(data), num_columns)]

            print("Sending JSON data:", rows)
            json_data = json.dumps(rows)

            # Send JSON data to the API
            api_url = 'http://127.0.0.1:8000/fetch/'
            try:
                response = requests.post(api_url, json=rows, headers={'Content-Type': 'application/json'})

                # Check if the request was successful (status code 200)
                if response.status_code == 200:
                    # Display the API response in json_display.html
                    api_response = response.json()
                    print("API Response:", api_response)
                    return render(request, 'json_display.html', {'json_data': json.dumps(api_response), 'api_response': api_response})
                else:
                    # If the request was not successful, handle the error accordingly
                    print(f'API request failed with status code {response.status_code}')
                    return render(request, 'error.html', {'error_message': f'API request failed with status code {response.status_code}'})
            except requests.RequestException as e:
                # Handle other request-related exceptions
                print(f'Request failed: {str(e)}')
                return render(request, 'error.html', {'error_message': f'Request failed: {str(e)}'})

    return render(request, 'upload_excel.html')
# def save_excel_data(request):
#     if request.method == "POST":
#         if "save_button" in request.POST:
#             # Example JSON array
#             json_array = [
#                 {
#                     "productName": "Fivestar",
#                     "weight": "",
#                     "mrp": "350",
#                     "sellPrice": "",
#                     "hsnCode": "",
#                     "gstPercent": "",
#                     "productCategory": ""
#                 },
#                 {
#                     "productName": "Gulabjamun",
#                     "weight": "500g",
#                     "mrp": "",
#                     "sellPrice": "",
#                     "hsnCode": "",
#                     "gstPercent": "",
#                     "productCategory": ""
#                 }
#             ]

#             # Send JSON data to the API
#             api_url = 'http://127.0.0.1:8000/fetch/'
#             try:
#                 response = requests.post(api_url, json=json_array, headers={'Content-Type': 'application/json'})

#                 # Check if the request was successful (status code 200)
#                 if response.status_code == 200:
#                     # Display the API response in json_display.html
#                     api_response = response.json()
#                     return render(request, 'json_display.html', {'json_data': json.dumps(json_array), 'api_response': api_response})
#                 else:
#                     # If the request was not successful, handle the error accordingly
#                     return render(request, 'error.html', {'error_message': f'API request failed with status code {response.status_code}'})
#             except requests.RequestException as e:
#                 # Handle other request-related exceptions
#                 return render(request, 'error.html', {'error_message': f'Request failed: {str(e)}'})

#     return render(request, 'upload_excel.html')
def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                next_url = request.POST.get('next', 'upload_excel')
                return redirect(next_url)

    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})

class CustomLoginView(LoginView):
    template_name = 'login.html'  # Your login template
    success_url = reverse_lazy('upload_excel')  # Redirect to 'upload_excel' upon successful login

# Views
@login_required
def home(request):
    # Display a message if the user was redirected here due to authentication failure
    if messages.get_messages(request):
        messages.error(request, 'Invalid username or password')

    return render(request, "login.html", {})
   
 
def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username = username, password = password)
            login(request, user)
            return redirect('upload_excel')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})